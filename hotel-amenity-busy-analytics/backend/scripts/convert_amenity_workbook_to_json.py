#!/usr/bin/env python3
"""Convert the amenity workbook into canonical JSON.

The generated JSON keeps every source column, normalizes field names to snake_case,
and stores lightweight schema metadata alongside all workbook records.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import date, datetime, time
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DEFAULT_SOURCE = Path(__file__).resolve().parents[1] / "data" / "hotel_amenity_large_dataset_60days_weather_traffic.xlsx"
DEFAULT_OUTPUT = Path(__file__).resolve().parents[1] / "data" / "hotel_amenity_large_dataset_60days_weather_traffic.canonical.json"


def to_snake_case(value: str) -> str:
    value = re.sub(r"(.)([A-Z][a-z]+)", r"\1_\2", value)
    value = re.sub(r"([a-z0-9])([A-Z])", r"\1_\2", value)
    value = re.sub(r"[^A-Za-z0-9]+", "_", value)
    return value.strip("_").lower()


def normalize_value(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, str):
        stripped = value.strip()
        lowered = stripped.lower()
        if lowered == "true":
            return True
        if lowered == "false":
            return False
        return stripped
    return value


def infer_type(values: list[Any]) -> str:
    non_null = [value for value in values if value is not None]
    if not non_null:
        return "null"
    if all(isinstance(value, bool) for value in non_null):
        return "boolean"
    if all(isinstance(value, int) and not isinstance(value, bool) for value in non_null):
        return "integer"
    if all(isinstance(value, (int, float)) and not isinstance(value, bool) for value in non_null):
        return "number"
    return "string"


def convert(source: Path, output: Path) -> dict[str, Any]:
    workbook = load_workbook(source, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    source_headers = [str(value) if value is not None else "" for value in next(rows)]
    canonical_fields = [to_snake_case(header) for header in source_headers]

    records: list[dict[str, Any]] = []
    samples: dict[str, list[Any]] = {field: [] for field in canonical_fields}

    for row in rows:
        record: dict[str, Any] = {}
        for field, value in zip(canonical_fields, row):
            normalized = normalize_value(value)
            record[field] = normalized
            if len(samples[field]) < 250:
                samples[field].append(normalized)
        records.append(record)

    schema = [
        {
            "name": field,
            "source_name": source_name,
            "type": infer_type(samples[field]),
        }
        for source_name, field in zip(source_headers, canonical_fields)
    ]

    payload = {
        "metadata": {
            "source_file": source.name,
            "source_sheet": worksheet.title,
            "row_count": len(records),
            "field_count": len(canonical_fields),
            "canonical_format": "snake_case_fields_with_schema_v1",
        },
        "schema": schema,
        "records": records,
    }

    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=False, separators=(",", ":")) + "\n")
    return payload["metadata"]


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()

    metadata = convert(args.source, args.output)
    print(json.dumps(metadata, indent=2))


if __name__ == "__main__":
    main()
