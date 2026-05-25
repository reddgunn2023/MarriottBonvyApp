"""Runtime CSV datasets and event logs for amenity/service verification.

When the user-provided large CSV exists locally, it is normalized into per-amenity
CSV files. In cloud/dev environments where that absolute path is unavailable, the
service generates deterministic one-month fallback CSVs so the app remains runnable.
"""

import csv
import os
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

from data.mock_slots import AMENITY_COLLECTIONS, TIME_SLOTS
from services.score_service import calculate_busy_score, calculate_demand_score

REPO_ROOT = Path(__file__).resolve().parents[3]
SOURCE_DATASET_CANDIDATES = [
    REPO_ROOT / "hotel-amenity-busy-analytics/backend/data/hotel_amenity_large_dataset_60days_weather_traffic.xlsx",
    REPO_ROOT / "hotel-amenity-busy-analytics/backend/src/data/hotel_amenity_large_dataset_60days_weather_traffic.xlsx",
    REPO_ROOT / "src/data/hotel_amenity_large_dataset_60days_weather_traffic.xlsx",
    REPO_ROOT / "src/data/hotel_amenity_large_dataset_60days_weather_traffic.xslx",
    REPO_ROOT / "src/data/hotel_amenity_large_dataset_60days.csv",
    REPO_ROOT / "hotel-amenity-busy-analytics/src/data/hotel_amenity_large_dataset_60days.csv",
    Path("/Users/sgunn825/Documents/hotel_amenity_large_dataset_60days_weather_traffic.xlsx"),
    Path("/Users/sgunn825/Documents/hotel_amenity_large_dataset_60days.csv"),
]


def _source_dataset() -> Path:
    override = os.environ.get("HOTEL_AMENITY_DATASET_PATH")
    if override:
        return Path(override)
    for candidate in SOURCE_DATASET_CANDIDATES:
        if candidate.exists():
            return candidate
    return SOURCE_DATASET_CANDIDATES[0]


SOURCE_DATASET = _source_dataset()
DATASET_DIR = Path(os.environ.get("HOTEL_AMENITY_DATASET_DIR", "/tmp/hotel-amenity-busy-analytics-datasets"))
EVENT_LOG = DATASET_DIR / "reservation_event_log.csv"
_SOURCE_ROWS_CACHE: list[dict] | None = None
DATASET_FIELDS = [
    "date",
    "time_slot",
    "property_id",
    "hotel_name",
    "city",
    "state",
    "amenity_id",
    "amenity_name",
    "category",
    "capacity",
    "booked",
    "available",
    "waitlist",
    "busy_score",
    "demand_score",
    "weather_condition",
    "traffic_condition",
    "weather_score",
    "traffic_score",
    "forecast_score",
    "last_guest_id",
    "reserved_guest_id",
    "cancelled_guest_id",
    "waitlist_guest_id",
    "last_event",
    "updated_at",
]
EVENT_FIELDS = [
    "timestamp",
    "event_type",
    "property_id",
    "guest_id",
    "slot_id",
    "amenity_id",
    "amenity_name",
    "date",
    "time_slot",
    "booked",
    "available",
    "waitlist",
    "busy_score",
    "demand_score",
    "weather_condition",
    "traffic_condition",
    "forecast_score",
    "message",
]

AMENITY_TYPE_ALIASES = {
    "breakfast": "free-breakfast",
    "freebreakfast": "free-breakfast",
    "evcharging": "ev-charging",
    "ev_charging": "ev-charging",
    "tennis": "tennis",
    "cabanas": "cabanas",
    "cabana": "cabanas",
    "restaurants": "restaurants",
    "restaurant": "restaurants",
    "pool": "pool",
    "spa": "spa",
    "golf": "golf",
    "fitness": "fitness-center",
    "fitnesscenter": "fitness-center",
    "whirlpool": "whirlpool-onsite",
    "hottub": "whirlpool-onsite",
}


def _amenity_from_source(amenity_type: str, service_name: str = "") -> dict | None:
    service_key = service_name.lower()
    # Service names are more precise than broad workbook amenityType values.
    if "breakfast" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "free-breakfast")
    if "lobby bar" in service_key or "bar" in service_key or "lounge" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "lounges")
    if "cabana" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "cabanas")
    if "tennis" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "tennis")
    if "charging" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "ev-charging")
    if "pool" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "pool")
    if "spa" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "spa")
    if "golf" in service_key:
        return next(item for item in AMENITY_COLLECTIONS if item["id"] == "golf")

    key = _norm_key(amenity_type).replace("_", "")
    alias_id = AMENITY_TYPE_ALIASES.get(key) or AMENITY_TYPE_ALIASES.get(_slug(amenity_type))
    for item in AMENITY_COLLECTIONS:
        normalized_collection = _norm_key(item.get("collection", "")).replace("_", "")
        if item["id"] == alias_id or normalized_collection == key or item["name"].lower() == amenity_type.lower():
            return item
    return None


def _norm_key(key: str) -> str:
    return key.strip().lower().replace(" ", "_").replace("-", "_")


def _get(row: dict, *aliases: str, default: str = "") -> str:
    normalized = {_norm_key(k): v for k, v in row.items()}
    for alias in aliases:
        value = normalized.get(_norm_key(alias))
        if value not in (None, ""):
            return str(value).strip()
    return default


def _safe_int(value: str, default: int = 0) -> int:
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default


def _safe_float(value: str, default: float = 0.0) -> float:
    try:
        return float(str(value).strip())
    except (TypeError, ValueError):
        return default


def _slug(value: str) -> str:
    return value.strip().lower().replace("&", "and").replace(" ", "-")


def weather_score(condition: str) -> float:
    value = condition.strip().lower()
    if value in {"severe", "storm", "hurricane", "extreme"}:
        return 1.0
    if value in {"rain", "heat", "snow", "wind", "hot"}:
        return 0.65
    if value in {"cloudy", "overcast", "moderate"}:
        return 0.35
    return 0.1


def traffic_score(condition: str) -> float:
    value = condition.strip().lower()
    if value in {"severe", "heavy", "high", "congested"}:
        return 1.0
    if value in {"moderate", "medium"}:
        return 0.55
    if value in {"light", "low"}:
        return 0.15
    return 0.25


def forecast_score(busy: float, demand: float, weather: float, traffic: float) -> float:
    return round(min(1.0, busy * 0.52 + min(demand, 1.5) * 0.24 + weather * 0.12 + traffic * 0.12), 2)


def weather_condition(day: date, time_index: int, amenity_name: str) -> str:
    """Deterministic dummy weather condition used by fallback CSV fixtures."""
    if day.day in {5, 12, 19, 26} and time_index in {2, 3, 4}:
        return "severe"
    if amenity_name in {"Pool", "Golf"} and time_index in {3, 4} and day.weekday() >= 5:
        return "heat"
    if day.day % 7 == 0:
        return "rain"
    return "clear"


def traffic_condition(day: date, time_index: int) -> str:
    if time_index in {1, 5} and day.weekday() < 5:
        return "heavy"
    if time_index in {2, 3, 4}:
        return "moderate"
    return "light"


def _dataset_path(amenity_id: str) -> Path:
    return DATASET_DIR / f"{amenity_id}.csv"


def _row_with_scores(row: dict) -> dict:
    capacity = max(_safe_int(row.get("capacity", 0)), 1)
    booked = min(_safe_int(row.get("booked", 0)), capacity)
    waitlist = max(_safe_int(row.get("waitlist", 0)), 0)
    available = max(_safe_int(row.get("available", capacity - booked), capacity - booked), 0)
    busy = _safe_float(row.get("busy_score", ""), calculate_busy_score(booked, capacity))
    demand = _safe_float(row.get("demand_score", ""), calculate_demand_score(booked, capacity, waitlist))
    w_score = _safe_float(row.get("weather_score", ""), weather_score(row.get("weather_condition", "clear")))
    t_score = _safe_float(row.get("traffic_score", ""), traffic_score(row.get("traffic_condition", "light")))
    return {
        **row,
        "capacity": str(capacity),
        "booked": str(booked),
        "available": str(available),
        "waitlist": str(waitlist),
        "busy_score": str(busy),
        "demand_score": str(demand),
        "weather_score": str(w_score),
        "traffic_score": str(t_score),
        "forecast_score": str(forecast_score(busy, demand, w_score, t_score)),
    }


def _base_rows(amenity: dict, start: date, days: int) -> list[dict]:
    rows = []
    for offset in range(days):
        current = start + timedelta(days=offset)
        for idx, time_slot in enumerate(TIME_SLOTS):
            capacity = amenity["capacity"]
            occupancy = ((idx + 1) * (offset + 3) + len(amenity["id"])) % (capacity + 1)
            booked = min(capacity, occupancy)
            available = max(capacity - booked, 0)
            waitlist = 1 if available == 0 else 0
            rows.append(
                _row_with_scores(
                    {
                        "date": current.isoformat(),
                        "time_slot": time_slot,
                        "property_id": "MARRIOTT101",
                        "hotel_name": "Residence Inn Anaheim Resort",
                        "city": "Anaheim",
                        "state": "California",
                        "amenity_id": amenity["id"],
                        "amenity_name": amenity["name"],
                        "category": amenity["category"],
                        "capacity": str(capacity),
                        "booked": str(booked),
                        "available": str(available),
                        "waitlist": str(waitlist),
                        "weather_condition": weather_condition(current, idx, amenity["name"]),
                        "traffic_condition": traffic_condition(current, idx),
                        "last_guest_id": "",
                        "reserved_guest_id": "",
                        "cancelled_guest_id": "",
                        "waitlist_guest_id": "",
                        "last_event": "seed",
                        "updated_at": "",
                    }
                )
            )
    return rows


def _normalize_source_row(row: dict) -> dict | None:
    amenity_type = _get(row, "amenityType", "amenity_type", "amenity", "amenity_name")
    service_name = _get(row, "serviceName", "service_name", "service", default=amenity_type)
    amenity = _amenity_from_source(amenity_type, service_name)
    if not amenity:
        return None
    capacity = _safe_int(_get(row, "totalCapacity", "capacity", "max_capacity", "total_capacity", default=str(amenity["capacity"])), amenity["capacity"])
    booked = _safe_int(_get(row, "bookedCount", "booked", "reserved", "reservations", "current_occupancy", "occupied", default="0"), 0)
    waitlist = _safe_int(_get(row, "waitlistCount", "waitlist", "waiting", "waiting_count", "waiting_line", default="0"), 0)
    available = _safe_int(_get(row, "availableCount", "available", "open", "remaining", default=str(max(capacity - booked, 0))), max(capacity - booked, 0))
    date_value = _get(row, "date", "slot_date", "stay_date", default=date.today().isoformat())
    time_slot = _get(row, "time_slot", "timeSlot", "time", "slot", "period")
    if not time_slot:
        start = _get(row, "timeSlotStart", "time_slot_start", default=TIME_SLOTS[0].split("-")[0])
        end = _get(row, "timeSlotEnd", "time_slot_end", default=TIME_SLOTS[0].split("-")[1])
        time_slot = f"{start}-{end}"
    return _row_with_scores(
        {
            "date": date_value,
            "time_slot": time_slot,
            "property_id": _get(row, "propertyId", "property_id", "property", "hotel_id", default="MARRIOTT101"),
            "hotel_name": _get(row, "hotelName", "hotel_name", default="Residence Inn Anaheim Resort"),
            "city": _get(row, "city", default="Anaheim"),
            "state": _get(row, "state", default="California"),
            "amenity_id": amenity["id"],
            "amenity_name": amenity["name"],
            "category": amenity["category"],
            "capacity": str(capacity),
            "booked": str(booked),
            "available": str(available),
            "waitlist": str(waitlist),
            "busy_score": _get(row, "busyScore", "busy_score", default=""),
            "demand_score": _get(row, "demandScore", "demand_score", default=""),
            "weather_condition": _get(row, "weatherCondition", "weather_condition", "weather", default="clear"),
            "traffic_condition": _get(row, "trafficLevel", "traffic_condition", "traffic", default="light"),
            "weather_score": _get(row, "weatherSeverityScore", "weather_score", default=""),
            "traffic_score": _get(row, "trafficScore", "traffic_score", default=""),
            "forecast_score": _get(row, "predictionScore", "forecast_score", "futureBusy", default=""),
            "last_guest_id": _get(row, "last_guest_id", "guest_id", "user_id", default=""),
            "reserved_guest_id": _get(row, "reserved_guest_id", default=""),
            "cancelled_guest_id": _get(row, "cancelled_guest_id", default=""),
            "waitlist_guest_id": _get(row, "waitlist_guest_id", default=""),
            "last_event": _get(row, "last_event", default="source"),
            "updated_at": _get(row, "updated_at", "lastUpdatedAt", default=""),
        }
    )


def _source_rows() -> list[dict]:
    global _SOURCE_ROWS_CACHE
    if _SOURCE_ROWS_CACHE is not None:
        return _SOURCE_ROWS_CACHE
    if not SOURCE_DATASET.exists():
        _SOURCE_ROWS_CACHE = []
        return _SOURCE_ROWS_CACHE
    if SOURCE_DATASET.suffix.lower() in {".xlsx", ".xslx"}:
        from openpyxl import load_workbook

        workbook = load_workbook(SOURCE_DATASET, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value) if value is not None else "" for value in next(iterator)]
        _SOURCE_ROWS_CACHE = [dict(zip(headers, row)) for row in iterator]
        return _SOURCE_ROWS_CACHE
    with SOURCE_DATASET.open(newline="") as handle:
        _SOURCE_ROWS_CACHE = list(csv.DictReader(handle))
        return _SOURCE_ROWS_CACHE


def _seed_from_source() -> list[str]:
    grouped: dict[str, list[dict]] = {}
    for row in _source_rows():
        normalized = _normalize_source_row(row)
        if normalized:
            grouped.setdefault(normalized["amenity_id"], []).append(normalized)
    created = []
    for amenity_id, rows in grouped.items():
        path = _dataset_path(amenity_id)
        with path.open("w", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=DATASET_FIELDS)
            writer.writeheader()
            writer.writerows(rows)
        created.append(str(path))
    return created


def dataset_properties() -> list[dict]:
    """Return distinct properties discovered from the source workbook/CSV."""
    properties: dict[str, dict] = {}
    for row in _source_rows():
        property_id = _get(row, "propertyId", "property_id", "property", default="")
        if not property_id or property_id in properties:
            continue
        hotel_name = _get(row, "hotelName", "hotel_name", default=property_id)
        properties[property_id] = {
            "id": property_id,
            "name": hotel_name,
            "location": ", ".join(part for part in [_get(row, "city"), _get(row, "state")] if part),
            "amenity_ids": [item["id"] for item in dataset_property_amenities(property_id)] if DATASET_DIR.exists() else [item["id"] for item in AMENITY_COLLECTIONS],
            "services": ["Digital Check In", "Mobile Key", "Service Request", "Wake-Up Calls", "Laundry", "Dry Cleaning Service", "Gift Shop"],
        }
    return list(properties.values())


def _source_property_ids() -> list[str]:
    ids: list[str] = []
    seen: set[str] = set()
    for row in _source_rows():
        property_id = _get(row, "propertyId", "property_id", "property", default="")
        if property_id and property_id not in seen:
            seen.add(property_id)
            ids.append(property_id)
    return ids


def canonical_property_id(property_id: str | None) -> str:
    """Map legacy/unknown property ids to the first workbook property when active."""
    if not SOURCE_DATASET.exists():
        return property_id or "prop-001"
    requested = property_id or ""
    property_ids = _source_property_ids()
    if requested in property_ids:
        return requested
    if requested.lower() in {"prop-001", "default", ""} and property_ids:
        return property_ids[0]
    return requested


def dataset_property_amenities(property_id: str) -> list[dict]:
    """Return only amenities/services present in the workbook/runtime rows for a property."""
    property_id = canonical_property_id(property_id)
    seed_csv_datasets()
    amenity_ids: set[str] = set()
    for path in sorted(DATASET_DIR.glob("*.csv")):
        if path.name == EVENT_LOG.name:
            continue
        for row in _read_rows(path):
            if row.get("property_id") == property_id:
                amenity_ids.add(row.get("amenity_id", ""))
                break
    by_id = {item["id"]: item for item in AMENITY_COLLECTIONS}
    preferred = ["free-breakfast", "pool", "fitness-center", "lounges"]
    ordered_ids = [item for item in preferred if item in amenity_ids]
    ordered_ids.extend(sorted(amenity_ids - set(ordered_ids)))
    return [{**by_id[item]} for item in ordered_ids if item in by_id]


def seed_csv_datasets(start: date | None = None, days: int = 30) -> dict:
    """Create per-amenity CSV datasets if they do not exist."""
    DATASET_DIR.mkdir(parents=True, exist_ok=True)
    created = []
    source_used = False
    if SOURCE_DATASET.exists() and not any(DATASET_DIR.glob("*.csv")):
        created.extend(_seed_from_source())
        source_used = True

    if not SOURCE_DATASET.exists():
        start = start or date.today().replace(day=1)
        for amenity in AMENITY_COLLECTIONS:
            path = _dataset_path(amenity["id"])
            if not path.exists():
                with path.open("w", newline="") as handle:
                    writer = csv.DictWriter(handle, fieldnames=DATASET_FIELDS)
                    writer.writeheader()
                    writer.writerows(_base_rows(amenity, start, days))
                created.append(str(path))

    if not EVENT_LOG.exists():
        with EVENT_LOG.open("w", newline="") as handle:
            csv.DictWriter(handle, fieldnames=EVENT_FIELDS).writeheader()

    return {
        "source_dataset": str(SOURCE_DATASET),
        "source_dataset_found": SOURCE_DATASET.exists(),
        "source_dataset_used": SOURCE_DATASET.exists(),
        "dataset_dir": str(DATASET_DIR),
        "event_log": str(EVENT_LOG),
        "created": created,
        "datasets": [str(path) for path in sorted(DATASET_DIR.glob("*.csv")) if path.name != EVENT_LOG.name],
    }


def _read_rows(path: Path) -> list[dict]:
    with path.open(newline="") as handle:
        return list(csv.DictReader(handle))


def _write_rows(path: Path, rows: list[dict]) -> None:
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=DATASET_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def update_dataset_for_event(slot: dict, event_type: str, guest_id: str, message: str) -> None:
    """Update the matching CSV row for the event's property/date/time period and log it."""
    seed_csv_datasets()
    path = _dataset_path(slot["amenityId"])
    rows = _read_rows(path)
    updated_at = datetime.now(timezone.utc).isoformat()
    matched = False
    for row in rows:
        same_period = row["date"] == slot["date"] and row["time_slot"] == slot["timeSlot"]
        same_property = row.get("property_id", slot["propertyId"]) == slot["propertyId"]
        if same_period and same_property:
            row.update(
                _row_with_scores(
                    {
                        **row,
                        "booked": str(slot["booked"]),
                        "available": str(slot["available"]),
                        "waitlist": str(slot.get("waitlist", 0)),
                        "weather_condition": slot.get("weatherCondition", row.get("weather_condition", "clear")),
                        "traffic_condition": slot.get("trafficCondition", row.get("traffic_condition", "light")),
                        "last_guest_id": guest_id,
                        "reserved_guest_id": guest_id if event_type == "RESERVE" else row.get("reserved_guest_id", ""),
                        "cancelled_guest_id": guest_id if event_type == "CANCEL" else row.get("cancelled_guest_id", ""),
                        "waitlist_guest_id": guest_id if event_type == "WAITLIST" else row.get("waitlist_guest_id", ""),
                        "last_event": event_type,
                        "updated_at": updated_at,
                    }
                )
            )
            matched = True
            break
    if not matched:
        rows.append(
            _row_with_scores(
                {
                    "date": slot["date"],
                    "time_slot": slot["timeSlot"],
                    "property_id": slot["propertyId"],
                    "amenity_id": slot["amenityId"],
                    "amenity_name": slot["amenity"],
                    "category": slot.get("category", "Service"),
                    "capacity": str(slot["capacity"]),
                    "booked": str(slot["booked"]),
                    "available": str(slot["available"]),
                    "waitlist": str(slot.get("waitlist", 0)),
                    "weather_condition": slot.get("weatherCondition", "clear"),
                    "traffic_condition": slot.get("trafficCondition", "light"),
                    "last_guest_id": guest_id,
                    "reserved_guest_id": guest_id if event_type == "RESERVE" else "",
                    "cancelled_guest_id": guest_id if event_type == "CANCEL" else "",
                    "waitlist_guest_id": guest_id if event_type == "WAITLIST" else "",
                    "last_event": event_type,
                    "updated_at": updated_at,
                }
            )
        )
    _write_rows(path, rows)

    event_row_base = _row_with_scores(
        {
            "capacity": str(slot["capacity"]),
            "booked": str(slot["booked"]),
            "available": str(slot["available"]),
            "waitlist": str(slot.get("waitlist", 0)),
            "weather_condition": slot.get("weatherCondition", "clear"),
            "traffic_condition": slot.get("trafficCondition", "light"),
        }
    )
    with EVENT_LOG.open("a", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=EVENT_FIELDS)
        writer.writerow(
            {
                "timestamp": updated_at,
                "event_type": event_type,
                "property_id": slot["propertyId"],
                "guest_id": guest_id,
                "slot_id": slot["slotId"],
                "amenity_id": slot["amenityId"],
                "amenity_name": slot["amenity"],
                "date": slot["date"],
                "time_slot": slot["timeSlot"],
                "booked": slot["booked"],
                "available": slot["available"],
                "waitlist": slot.get("waitlist", 0),
                "busy_score": event_row_base["busy_score"],
                "demand_score": event_row_base["demand_score"],
                "weather_condition": event_row_base["weather_condition"],
                "traffic_condition": event_row_base["traffic_condition"],
                "forecast_score": event_row_base["forecast_score"],
                "message": message,
            }
        )


def _parse_time_minutes(value: str) -> int | None:
    try:
        hour, minute = value.strip().replace('.', ':').split(':')
        return int(hour) * 60 + int(minute)
    except (ValueError, AttributeError):
        return None


def _format_minutes(value: int) -> str:
    value = value % (24 * 60)
    return f"{value // 60:02d}:{value % 60:02d}"


def _expand_row_to_30_minute_rows(row: dict) -> list[dict]:
    slot = row.get("time_slot", TIME_SLOTS[0])
    if '-' not in slot:
        return [row]
    start_text, end_text = slot.split('-', 1)
    start = _parse_time_minutes(start_text)
    end = _parse_time_minutes(end_text)
    if start is None or end is None:
        return [row]
    if end <= start:
        end += 24 * 60
    if end - start <= 30:
        normalized = {**row, "time_slot": f"{_format_minutes(start)}-{_format_minutes(start + 30)}"}
        return [normalized]
    rows = []
    current = start
    while current < end:
        rows.append({**row, "time_slot": f"{_format_minutes(current)}-{_format_minutes(current + 30)}"})
        current += 30
    return rows


def _dataset_row_to_slot(row: dict, slot_index: int) -> dict:
    capacity = max(_safe_int(row.get("capacity", 1)), 1)
    booked = min(_safe_int(row.get("booked", 0)), capacity)
    waitlist = max(_safe_int(row.get("waitlist", 0)), 0)
    available = max(_safe_int(row.get("available", capacity - booked), capacity - booked), 0)
    time_slot = row.get("time_slot", TIME_SLOTS[0])
    try:
        time_index = TIME_SLOTS.index(time_slot)
    except ValueError:
        time_index = slot_index % len(TIME_SLOTS)
    amenity_id = row.get("amenity_id", _slug(row.get("amenity_name", "amenity")))
    property_id = row.get("property_id", "prop-001")
    slot_id = f"{property_id}-{amenity_id}-{row.get('date', date.today().isoformat())}-{time_index}"
    normalized = _row_with_scores(
        {
            **row,
            "capacity": str(capacity),
            "booked": str(booked),
            "available": str(available),
            "waitlist": str(waitlist),
        }
    )
    return {
        "slotId": slot_id,
        "propertyId": property_id,
        "amenityId": amenity_id,
        "amenity": row.get("amenity_name", amenity_id),
        "category": row.get("category", "Service"),
        "serviceType": row.get("service_type", "reservation"),
        "date": row.get("date", date.today().isoformat()),
        "timeSlot": time_slot,
        "timeIndex": time_index,
        "capacity": capacity,
        "booked": booked,
        "available": available,
        "waitlist": waitlist,
        "waitlistGuests": [],
        "reservedGuests": [],
        "weatherCondition": normalized.get("weather_condition", "clear"),
        "trafficCondition": normalized.get("traffic_condition", "light"),
        "busyScore": float(normalized.get("busy_score", 0)),
        "demandScore": float(normalized.get("demand_score", 0)),
        "futureBusy": float(normalized.get("forecast_score", 0)),
        "status": "FULL" if available <= 0 else "AVAILABLE",
    }


def get_slots_from_datasets(
    property_id: str = "prop-001",
    check_in: str | None = None,
    check_out: str | None = None,
) -> list[dict]:
    """Return bookable slots from the normalized CSV datasets."""
    property_id = canonical_property_id(property_id)
    seed_csv_datasets()
    start = date.fromisoformat(check_in) if check_in else None
    end = date.fromisoformat(check_out) if check_out else None
    slots = []
    for path in sorted(DATASET_DIR.glob("*.csv")):
        if path.name == EVENT_LOG.name:
            continue
        slot_index = 0
        for row in _read_rows(path):
            for expanded_row in _expand_row_to_30_minute_rows(row):
                row_property = expanded_row.get("property_id", property_id)
                if row_property != property_id:
                    continue
                row_date = expanded_row.get("date", "")
                if start and row_date and date.fromisoformat(row_date) < start:
                    continue
                if end and row_date and date.fromisoformat(row_date) >= end:
                    continue
                slots.append(_dataset_row_to_slot(expanded_row, slot_index))
                slot_index += 1
    return slots


def recent_events(limit: int = 25) -> list[dict]:
    seed_csv_datasets()
    rows = _read_rows(EVENT_LOG)
    return rows[-limit:]
